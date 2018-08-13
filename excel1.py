from openpyxl import load_workbook


def copy_excel():
    init_row = 5        # 在目标文件中复制的起始行数

    # 加载目标文件
    target = load_workbook('target.xlsx')
    source_file_dir = target["1111"]["B2"].value
    # 加载数据文件
    source = load_workbook(filename=source_file_dir)

    for names in target.sheetnames:
        key = target[names]["B3"].value
        if key == "sh":
            source_sheet = source["sh"]
            max_column = 9               # A:I

            # 清空原target文件内容
            for m in range(init_row, target[names].max_row+1):
                for n in range(1, max_column + 1):
                    target[names].cell(row=m, column=n).value = None

            # 复制source的内容到target文件中
            source_cells = source_sheet['A5': 'I' + str(source_sheet.max_row)]
            for i in range(source_sheet.max_row - init_row+1):
                for j in range(max_column):
                    if source_cells[i][j].value is None:
                        continue
                    else:
                        target[names].cell(row=i + init_row, column=j+1).value = source_cells[i][j].value

        elif key == "wy":
            source_sheet = source["wy"]
            max_column = 11                # A:K

            # 清空原target文件内容
            for m in range(init_row, target[names].max_row + 1):
                for n in range(1, max_column + 1):
                    target[names].cell(row=m, column=n).value = None

            # 复制source的内容到target文件中
            source_cells = source_sheet['A5': 'K'+str(source_sheet.max_row)]
            for i in range(source_sheet.max_row - init_row + 1):
                for j in range(max_column):
                    if source_cells[i][j].value is None:
                        continue
                    else:
                        target[names].cell(row=i + init_row, column=j + 1).value = source_cells[i][j].value
        else:
            continue

    target.save(filename='target.xlsx')


if __name__ == "__main__":
    copy_excel()

